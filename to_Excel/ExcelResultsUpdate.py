import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

#useful functions
def add_headers(wb, par_list) :
    for name in wb.sheetnames :
        ws = wb[name]
        i = 1
        for x in par_list :
            ws.cell(column = i, row = 1, value = x)
            i += 1

def is_sheet_name_present(wb, x) :
    for name in wb :
        if name == "W"+str(x[0]):
            sname = 1
        else:
            sname = 0
    return sname

def find_row(starting_row, col, value) :
    row_index = -1
    k = starting_row
    for cell in col :
        if cell.value == value : #looks for fluence zone and gets the correct line index
            row_index = k
            break
        k += 1
    return row_index

def find_bias_row(starting_row, col, value) :
    row_index = -1
    k = starting_row
    for cell in col :
        if value == cell.value  :
            print(f"The bias point @ {value} V was already present, it was added anyway")
            row_index = k
            break
        elif value < int(cell.value) :
            row_index = k
            break
        k += 1
    return row_index

#future header of the excel file
par_list = [
'Wafer',
'Fluence [neq/cm2]',
'Temp',
'Bias',
'Amp',
'Area [pWb]',
'Charge [fC]',
'gain',
'dVdt [mV/ns]',
'rms [mV]',
'risetime [ps]',
'SNR',
'jitter [ps]',
'Landau [ps]',
'Resolution [ps]',
'Total resolution [ps]',
'Amp cuts',
'Area cuts',
'MPV (amp)',
'Trigger resolution [ps]'
]

df = pd.read_csv('data_sample.txt', delimiter = ' ', header = None, names = par_list)
data = df.to_numpy()

#print(data)

wb = load_workbook(filename = 'LogbookUFSD3.2.xlsx')

#adds the header to each sheet
#add_headers(wb, par_list)

ws_name = "W"+str(data[0][0])
ws = wb[ws_name]   #opens correct sheet, assumes only data from a certain wafer are present

flue_col = ws['B']
temp_col    = ws['C']
bias_col    = ws['D']

flue_row = find_row(1, flue_col, data[0][1]) #assumes only one fluence per data file
temp_row = find_row(flue_row, temp_col, data[0][2]) #assumes only one temperature per data file

i = 1
for data_row in data :
    if i == 1 :
        bias_row = find_bias_row(temp_row, bias_col, data_row[3])
    else :
        bias_row = find_bias_row(i       , bias_col, data_row[3])
    print(f"bias row is {bias_row}")
    i = bias_row + 1
    ws.insert_rows(i)
    j = 1
    for x in data_row :
        ws.cell(column = j, row = i, value = x)
        j += 1

wb.save(filename='LogbookUFSD3.2.xlsx')
